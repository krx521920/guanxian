#!/usr/bin/env python3
"""Fill the downloaded official member template for one isolated E2E member."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook


HEADERS = [
    "企业名称*", "统一社会信用代码", "企业分类*", "联系地址", "联系人", "联系电话",
    "联系邮箱", "企业简介", "核心能力（用；分隔）", "产品（用；分隔）", "服务（用；分隔）",
    "应用场景（用；分隔）", "合作需求（用；分隔）", "可见范围",
]


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--template", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--name", required=True)
    parser.add_argument("--credit-code", required=True)
    args = parser.parse_args()

    workbook = load_workbook(args.template)
    sheet = workbook["会员资料"]
    actual_headers = [sheet.cell(1, column).value for column in range(1, len(HEADERS) + 1)]
    if actual_headers != HEADERS:
        raise ValueError("downloaded survey template headers changed; update the E2E fixture explicitly")
    values = [
        args.name, args.credit_code, "E2E测试企业", "北京市海淀区", "端到端联系人", "13800000009",
        "e2e-member@invalid.example", "由真实调查模板导入的隔离端到端企业。", "管线检测；数据治理",
        "检测终端", "现场检测", "燃气；供热", "寻找项目合作方", "MEMBERS",
    ]
    for column, value in enumerate(values, start=1):
        sheet.cell(2, column).value = value
    workbook["提交信息"].cell(2, 2).value = "E2E自动化验收提交单位"
    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
